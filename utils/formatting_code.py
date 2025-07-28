from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import datetime

import warnings
warnings.simplefilter(action='ignore', category=UserWarning)

class FormatExcel:
    @staticmethod
    def formatExcelFile(file_name: str):
        '''
            This is a general format for formatting the excel files.
        '''
        
        wb = load_workbook(file_name)
        
        # Borders
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        no_border = Border()  # No border style
        
        # Header styling
        header_fill = PatternFill(start_color='1256ff', end_color='1256ff', fill_type='solid')
        header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        # Columns to format as percentage
        percentage_columns = {"CTR", "Order Conversion Rate", "ACOS"}
        
        # Create a percentage style (Only for Display, Keeps Actual Values)
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = "0.0%"  # Display as percentage
        percentage_style.alignment = alignment_center
        percentage_style.border = thin_border
        
        # Create a number rounding style (Only for Display, Keeps Actual Values)
        number_style = NamedStyle(name="rounded_number")
        number_style.number_format = "0.0"  # Display rounded, keeps actual value
        number_style.alignment = alignment_center
        number_style.border = thin_border
        
        # Loop through all sheets
        for sheet in wb.worksheets:
            if sheet.max_row == 0 or sheet.max_column == 0:  # Skip empty sheets
                continue

            # Ensure First Row Height is Set to 23
            sheet.row_dimensions[1].height = 23  

            # Detect column types from the second row
            column_types = {}
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                header_value = str(sheet.cell(row=1, column=col).value).strip() if sheet.cell(row=1, column=col).value else ""
                
                if not header_value:  # If header is empty, skip formatting for this column
                    sheet.column_dimensions[col_letter].width = 10
                    for row in range(1, sheet.max_row + 1):
                        sheet.cell(row=row, column=col).border = no_border  # Remove borders
                    continue
                
                for row in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        if isinstance(cell_value, (int, float)):
                            column_types[col_letter] = "number"
                        elif isinstance(cell_value, (datetime.datetime, datetime.date)):
                            column_types[col_letter] = "datetime"
                        elif isinstance(cell_value, str):
                            column_types[col_letter] = "text"
                        break  # Stop checking once we identify the type

            # Apply formatting
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                max_width = 0
                header_value = str(sheet.cell(row=1, column=col).value).strip() if sheet.cell(row=1, column=col).value else ""
                
                if not header_value:
                    continue  # Skip formatting if the header is empty
                
                for row in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border  # Apply thin border
                    
                    # Apply header formatting
                    if row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = alignment_center
                    else:
                        cell_type = column_types.get(col_letter, "text")
                        
                        if cell_type == "number":
                            if isinstance(cell.value, float):
                                if header_value in percentage_columns:
                                    cell.style = percentage_style  # Display as percentage
                                else:
                                    cell.style = number_style  # Display rounded without modifying value
                            cell.alignment = alignment_center
                        else:
                            cell.alignment = alignment_left

                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)) + 2)  # Adjust width
                
                # Adjust column width
                final_width = 40 if max_width > 40 else max_width + 2
                sheet.column_dimensions[col_letter].width = final_width

        wb.save(file_name)
        print(f"{file_name} File Formatted Successfully")